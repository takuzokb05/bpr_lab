import json, re, html, subprocess, time, os, sys
import openpyxl

SRC = '/root/.claude/uploads/54b845d3-b34b-5735-ba7a-d98ad5a184da/29ed86ec-________URL__.xlsx'
OUT = '/tmp/claude-0/-home-user-bpr-lab/54b845d3-b34b-5735-ba7a-d98ad5a184da/scratchpad'
PAGES = os.path.join(OUT, 'pages')
UA = "Mozilla/5.0 (compatible; url-existence-check/1.0)"
ERR_TITLE_MARK = 'お探しのページは見つかりませんでした'

def fetch(url, dest):
    """returns dict or None on failure"""
    fmt = "%{http_code}\t%{num_redirects}\t%{url_effective}\t%{size_download}"
    p = subprocess.run(
        ["curl","-sSL","-o",dest,"-w",fmt,"-A",UA,"--max-time","45","--retry","0",url],
        capture_output=True, text=True)
    if p.returncode != 0 or not p.stdout.strip():
        return None
    code, nred, eff, size = p.stdout.strip().split("\t")
    body = open(dest, encoding='utf-8', errors='replace').read()
    t = re.search(r'<title[^>]*>(.*?)</title>', body, re.S)
    title = html.unescape(re.sub(r'\s+',' ', t.group(1))).strip() if t else ''
    return dict(code=int(code), redirects=int(nred), effective=eff, size=int(size),
                title=title, has_kw=('公示送達' in body),
                kw_count=body.count('公示送達'),
                err_page=(ERR_TITLE_MARK in title))

wb = openpyxl.load_workbook(SRC)
results = []
for ws in wb.worksheets:
    for r in range(6, 24):
        ward = ws.cell(r, 1).value
        url = ws.cell(r, 2).value
        dest = os.path.join(PAGES, f"{ws.title}_{r}.html")
        res = None
        attempts = 0
        for attempt in range(3):          # 初回 + リトライ2回
            attempts = attempt + 1
            res = fetch(url, dest)
            if res is not None and res['code'] not in (0, 429, 500, 502, 503, 504):
                break
            if attempt < 2:
                time.sleep(3)
        rec = dict(sheet=ws.title, row=r, ward=ward, url=url, attempts=attempts)
        rec.update(res if res else dict(code=None, error='fetch failed'))
        results.append(rec)
        print(json.dumps(rec, ensure_ascii=False), flush=True)
        time.sleep(1)

json.dump(results, open(os.path.join(OUT,'results.json'),'w'), ensure_ascii=False, indent=1)
