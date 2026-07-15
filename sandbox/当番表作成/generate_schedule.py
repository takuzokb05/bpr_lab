"""
2026年3月 当番表自動生成スクリプト v2

修正点:
- 同一時間帯に同一班から3人以上不可（2人まで許容→3人以上禁止に正確化）
- 連続営業日に同じ人が当番に入らない（中西以外）
"""

import random
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import datetime

# ========================================
# 職員データ
# ========================================
STAFF = [
    ("中西", "C"),
    ("前園", "C"),
    ("秋庭", "A"),
    ("池田", "D"),
    ("伊澤", "E"),
    ("遠藤（一）", "A"),
    ("遠藤（幹）", "B"),
    ("大泉", "D"),
    ("岡田", "E"),
    ("岡山", "C"),
    ("亀野", "A"),
    ("桐ヶ谷", "D"),
    ("小峰", "D"),
    ("齋藤", "D"),
    ("澤口", "E"),
    ("高澤", "A"),
    ("鈴木", "E"),
    ("須山", "B"),
    ("高橋", "E"),
    ("辻倉", "C"),
    ("中井", "A"),
    ("中田", "B"),
    ("中村", "B"),
    ("平木", "B"),
    ("平田", "C"),
    ("福田", "B"),
    ("本間", "E"),
    ("吉田", "B"),
    ("渡邊", "A"),
    ("本田", "C"),
    ("郡志", "C"),
    ("仲川", "C"),
    ("藤巻", "C"),
]

# 郡志は水曜定休日
WEDNESDAY_OFF = {"郡志"}

STAFF_DICT = {name: group for name, group in STAFF}
STAFF_NAMES = [name for name, _ in STAFF]
PHONE_CANDIDATES = [name for name in STAFF_NAMES if name != "中西"]

# ========================================
# グループ窓当番設定（23日以降）
# ========================================
GROUP_WINDOW_START_DAY = 23
GROUP_ORDER = ["A", "B", "C", "D", "E"]
GROUP_MEMBERS = defaultdict(list)
for _n, _g in STAFF:
    GROUP_MEMBERS[_g].append(_n)

# ========================================
# 2026年3月の営業日
# ========================================
HOLIDAYS_2026_MAR = [
    datetime.date(2026, 3, 20),  # 春分の日
]

def get_business_days(year, month):
    days = []
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        for i, day in enumerate(week):
            if day == 0:
                continue
            if i >= 5:  # 土(5)・日(6)
                continue
            d = datetime.date(year, month, day)
            if d in HOLIDAYS_2026_MAR:
                continue
            days.append(d)
    return days

BUSINESS_DAYS = get_business_days(2026, 3)
WEEKDAY_JP = ["月", "火", "水", "木", "金"]

# ========================================
# ヘルパー関数
# ========================================

def group_count_in_slot(members, group):
    return sum(1 for m in members if STAFF_DICT[m] == group)

def can_add_to_slot(candidate, current_members):
    """同一班が既に2人いたら追加不可（3人以上になるため）"""
    group = STAFF_DICT[candidate]
    return group_count_in_slot(current_members, group) < 2

# ========================================
# 当番表生成
# ========================================

def assign_schedule(seed=42):
    random.seed(seed)
    
    am_window_count = defaultdict(int)
    pm_window_count = defaultdict(int)
    phone_count = defaultdict(int)
    
    schedule = {}
    # 前日当番記録（連続禁止チェック用）
    prev_day_assigned = set()  # 前営業日に当番に出た人（中西除く）

    # 個人割り当て期間とグループ期間を分離
    individual_days = [d for d in BUSINESS_DAYS if d.day < GROUP_WINDOW_START_DAY]
    group_days = [d for d in BUSINESS_DAYS if d.day >= GROUP_WINDOW_START_DAY]

    for day in individual_days:
        weekday = day.weekday()  # 0=月, 1=火, 2=水

        # 水曜定休の職員を除外
        day_off = {s for s in WEDNESDAY_OFF if weekday == 2}

        # 連続禁止セット（中西は除外）
        rest_required = prev_day_assigned - {"中西"}

        # -----------------------------------------------
        # Step1: 午前窓（2名）
        # -----------------------------------------------
        am_window = []
        # 火・水は中西固定
        if weekday in [1, 2]:
            am_window = ["中西"]

        am_candidates = [
            s for s in STAFF_NAMES
            if s not in am_window
            and s not in rest_required  # 連続禁止
            and s not in day_off        # 定休日除外
            and can_add_to_slot(s, am_window)
        ]
        # 均等化：窓合計回数が少ない順、同数はランダム
        am_candidates.sort(key=lambda s: (am_window_count[s] + pm_window_count[s], random.random()))

        for cand in am_candidates:
            if len(am_window) >= 2:
                break
            if can_add_to_slot(cand, am_window):
                am_window.append(cand)

        if len(am_window) < 2:
            return None, None, None, None  # 失敗

        # -----------------------------------------------
        # Step2: 午後窓（2名）
        # -----------------------------------------------
        pm_candidates = [
            s for s in STAFF_NAMES
            if s not in am_window
            and s not in rest_required  # 連続禁止
            and s not in day_off        # 定休日除外
            and can_add_to_slot(s, [])
        ]
        pm_candidates.sort(key=lambda s: (am_window_count[s] + pm_window_count[s], random.random()))

        pm_window = []
        for cand in pm_candidates:
            if len(pm_window) >= 2:
                break
            if can_add_to_slot(cand, pm_window):
                pm_window.append(cand)

        if len(pm_window) < 2:
            return None, None, None, None

        # -----------------------------------------------
        # Step3: 電話当番（6名：午前3名＋午後3名）
        # -----------------------------------------------
        window_staff = set(am_window + pm_window)
        phone_pool = [
            s for s in PHONE_CANDIDATES
            if s not in window_staff
            and s not in rest_required  # 連続禁止
            and s not in day_off        # 定休日除外
        ]
        phone_pool.sort(key=lambda s: (phone_count[s], random.random()))

        phone = []
        for cand in phone_pool:
            if len(phone) >= 6:
                break
            if can_add_to_slot(cand, phone):
                phone.append(cand)

        if len(phone) < 6:
            return None, None, None, None

        # カウント更新
        for s in am_window:
            am_window_count[s] += 1
        for s in pm_window:
            pm_window_count[s] += 1
        for s in phone:
            phone_count[s] += 1

        schedule[day] = {
            "am_window": am_window,
            "pm_window": pm_window,
            "phone_am": phone[:3],
            "phone_pm": phone[3:],
        }

        # 翌日用：今日当番に出た人を記録（中西以外）
        today_assigned = set(am_window + pm_window + phone) - {"中西"}
        prev_day_assigned = today_assigned

        # phone変数はカウント更新後に分割済みなのでここでは全体setを使う

    # ===================================
    # 23日以降：班ローテーション窓当番
    # ===================================
    group_idx = 0
    for day in group_days:
        weekday = day.weekday()
        day_off = {s for s in WEDNESDAY_OFF if weekday == 2}
        rest_required = prev_day_assigned - {"中西"}

        current_group = GROUP_ORDER[group_idx % len(GROUP_ORDER)]
        group_idx += 1

        window_group_set = set(GROUP_MEMBERS[current_group])

        # 電話当番（6名）：窓当番班を除外
        phone_pool = [
            s for s in PHONE_CANDIDATES
            if s not in window_group_set
            and s not in rest_required
            and s not in day_off
        ]
        phone_pool.sort(key=lambda s: (phone_count[s], random.random()))

        phone = []
        for cand in phone_pool:
            if len(phone) >= 6:
                break
            if can_add_to_slot(cand, phone):
                phone.append(cand)

        if len(phone) < 6:
            return None, None, None, None

        for s in phone:
            phone_count[s] += 1

        schedule[day] = {
            "window_group": current_group,
            "phone_am": phone[:3],
            "phone_pm": phone[3:],
        }

        today_assigned = (window_group_set | set(phone)) - {"中西"}
        prev_day_assigned = today_assigned

        # phone変数はカウント更新後に分割済みなのでここでは全体setを使う

    return schedule, am_window_count, pm_window_count, phone_count

def try_assign(max_attempts=2000):
    for seed in range(max_attempts):
        result = assign_schedule(seed=seed)
        if result[0] is not None:
            print(f"成功 (seed={seed})")
            return result
    print("警告: 最大試行回数到達")
    return assign_schedule(seed=0)

# ========================================
# 検証
# ========================================

def verify_schedule(schedule):
    errors = []
    days_list = sorted(schedule.keys())
    prev_assigned = set()
    
    for i, day in enumerate(days_list):
        slots = schedule[day]

        phone_am = slots.get("phone_am", [])
        phone_pm = slots.get("phone_pm", [])
        phone_all = phone_am + phone_pm

        # 郡志の水曜チェック
        if day.weekday() == 2:
            for s in WEDNESDAY_OFF:
                if s in phone_all or s in slots.get("am_window", []) or s in slots.get("pm_window", []):
                    errors.append(f"{day}: {s} が水曜（定休日）に割り当て")

        if "window_group" in slots:
            # === グループ窓当番の日 ===
            group = slots["window_group"]
            window_group_set = set(GROUP_MEMBERS[group])

            # 電話枠の班チェック
            for slot_name, members in [("電話午前", phone_am), ("電話午後", phone_pm)]:
                for g in "ABCDE":
                    cnt = sum(1 for m in members if STAFF_DICT[m] == g)
                    if cnt >= 3:
                        errors.append(f"{day} {slot_name}: {g}班が{cnt}人")

            # 連続チェック（電話当番のみ）
            if i > 0:
                rest_check = prev_assigned - {"中西"}
                for s in rest_check:
                    if s in phone_all:
                        errors.append(f"{day}: {s} が連続営業日当番（電話）")

            prev_assigned = window_group_set | set(phone_all)
        else:
            # === 個人窓当番の日 ===
            am = slots["am_window"]
            pm = slots["pm_window"]

            all_assigned = am + pm + phone_all
            if len(all_assigned) != len(set(all_assigned)):
                dup = [x for x in set(all_assigned) if all_assigned.count(x) > 1]
                errors.append(f"{day}: 重複あり {dup}")

            for slot_name, members in [("午前窓", am), ("午後窓", pm), ("電話午前", phone_am), ("電話午後", phone_pm)]:
                for g in "ABCDE":
                    cnt = sum(1 for m in members if STAFF_DICT[m] == g)
                    if cnt >= 3:
                        errors.append(f"{day} {slot_name}: {g}班が{cnt}人 → {[m for m in members if STAFF_DICT[m]==g]}")

            if "中西" in phone_all:
                errors.append(f"{day}: 中西が電話当番")
            weekday = day.weekday()
            if weekday in [1, 2] and "中西" not in am:
                errors.append(f"{day}: 火・水なのに中西が午前窓にいない")

            if i > 0:
                rest_check = prev_assigned - {"中西"}
                for s in rest_check:
                    if s in all_assigned:
                        errors.append(f"{day}: {s} が連続営業日当番")

            prev_assigned = set(all_assigned)

    return errors

# ========================================
# Excel出力
# ========================================

GROUP_COLORS = {
    "A": "D7E8FF",
    "B": "D7FFD7",
    "C": "FFFFD7",
    "D": "FFE8D7",
    "E": "FFD7FF",
}

def get_fill(group):
    hex_color = GROUP_COLORS.get(group, "FFFFFF")
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def export_to_excel(schedule, am_count, pm_count, phone_count):
    wb = openpyxl.Workbook()

    # ========== シート1: メインカレンダー ==========
    ws = wb.active
    ws.title = "2026年3月 当番表"

    # タイトル
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "2026年3月 当番表"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ヘッダー
    headers = ["日付", "曜日", "午前窓①", "午前窓②", "午後窓①", "午後窓②",
               "電話(午前)①", "電話(午前)②", "電話(午前)③",
               "電話(午後)④", "電話(午後)⑤", "電話(午後)⑥"]
    col_widths = [10, 5, 10, 10, 10, 10, 12, 12, 12, 12, 12, 12]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    for day in BUSINESS_DAYS:
        weekday = day.weekday()
        weekday_str = WEEKDAY_JP[weekday]
        slots = schedule.get(day, {})
        is_group_day = "window_group" in slots

        phone_am = slots.get("phone_am", [])
        phone_pm = slots.get("phone_pm", [])

        if is_group_day:
            group = slots["window_group"]
            row_data = (
                [f"{day.month}/{day.day}", weekday_str]
                + [f"{group}班", ""]
                + [f"{group}班", ""]
                + phone_am + [""] * (3 - len(phone_am))
                + phone_pm + [""] * (3 - len(phone_pm))
            )
        else:
            am = slots.get("am_window", [])
            pm = slots.get("pm_window", [])
            row_data = (
                [f"{day.month}/{day.day}", weekday_str]
                + am + [""] * (2 - len(am))
                + pm + [""] * (2 - len(pm))
                + phone_am + [""] * (3 - len(phone_am))
                + phone_pm + [""] * (3 - len(phone_pm))
            )

        row_idx = 2 + BUSINESS_DAYS.index(day) + 1

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

            if col_idx == 2:  # 曜日
                if weekday == 1:
                    cell.font = Font(color="FF0000")
                elif weekday == 2:
                    cell.font = Font(color="0070C0")

            if col_idx >= 3 and value:
                if is_group_day and col_idx in [3, 5]:
                    # グループ窓当番セル：班色＋太字
                    cell.fill = get_fill(slots["window_group"])
                    cell.font = Font(bold=True)
                else:
                    staff_group = STAFF_DICT.get(value)
                    if staff_group:
                        cell.fill = get_fill(staff_group)
                        if value == "中西" and weekday in [1, 2] and col_idx in [3, 4]:
                            cell.font = Font(bold=True)

        ws.row_dimensions[row_idx].height = 18

    # 凡例（班色）
    legend_row = 2 + len(BUSINESS_DAYS) + 2
    ws.cell(row=legend_row, column=1, value="【班色凡例】").font = Font(bold=True)
    for i, (g, color) in enumerate(GROUP_COLORS.items()):
        c = ws.cell(row=legend_row, column=i+2, value=f"{g}班")
        c.fill = get_fill(g)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    # ========== シート2: 集計表 ==========
    ws2 = wb.create_sheet("集計")
    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "当番回数集計"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A1"].alignment = Alignment(horizontal="center")

    headers2 = ["名前", "班", "午前窓", "午後窓", "窓合計", "電話", "総当番数"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = 12

    for i, (name, group) in enumerate(sorted(STAFF, key=lambda x: x[1]), start=3):
        am_c = am_count[name]
        pm_c = pm_count[name]
        ph_c = phone_count[name]
        row_data = [name, group, am_c, pm_c, am_c + pm_c, ph_c, am_c + pm_c + ph_c]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
            if col_idx == 2:
                cell.fill = get_fill(group)

    # ========== シート3: 個人別一覧 ==========
    ws3 = wb.create_sheet("個人別")
    ws3.merge_cells("A1:C1")
    ws3["A1"].value = "個人別 当番一覧"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3["A1"].alignment = Alignment(horizontal="center")

    for col_idx, (h, w) in enumerate(zip(["名前", "班", "当番日程（種別）"], [12, 5, 100]), start=1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = thin_border()
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    for i, (name, group) in enumerate(sorted(STAFF, key=lambda x: (x[1], x[0])), start=3):
        duties = []
        for day in sorted(schedule.keys()):
            wd = WEEKDAY_JP[day.weekday()]
            label = f"{day.month}/{day.day}({wd})"
            slots = schedule[day]
            if "window_group" in slots:
                # グループ窓当番の日
                if STAFF_DICT[name] == slots["window_group"]:
                    duties.append(f"{label}:窓({slots['window_group']}班)")
                if name in slots.get("phone_am", []):
                    duties.append(f"{label}:電話(午前)")
                if name in slots.get("phone_pm", []):
                    duties.append(f"{label}:電話(午後)")
            else:
                if name in slots.get("am_window", []):
                    duties.append(f"{label}:午前窓")
                if name in slots.get("pm_window", []):
                    duties.append(f"{label}:午後窓")
                if name in slots.get("phone_am", []):
                    duties.append(f"{label}:電話(午前)")
                if name in slots.get("phone_pm", []):
                    duties.append(f"{label}:電話(午後)")
        ws3.cell(row=i, column=1, value=name).border = thin_border()
        g_cell = ws3.cell(row=i, column=2, value=group)
        g_cell.fill = get_fill(group)
        g_cell.border = thin_border()
        d_cell = ws3.cell(row=i, column=3, value="  /  ".join(duties))
        d_cell.border = thin_border()

    output_path = r"c:\Users\takuz\プロジェクト\bpr_lab\sandbox\当番表作成\2026年3月_当番表.xlsx"
    wb.save(output_path)
    return output_path

# ========================================
# メイン
# ========================================

if __name__ == "__main__":
    print("=== 2026年3月 当番表生成 ===")
    print(f"営業日数: {len(BUSINESS_DAYS)}日")

    schedule, am_count, pm_count, phone_count = try_assign()

    if schedule is None:
        print("ERROR: スケジュール生成失敗")
        exit(1)

    print("\n=== 検証 ===")
    errors = verify_schedule(schedule)
    if errors:
        print("エラーあり:")
        for e in errors:
            print(f"  x {e}")
    else:
        print("OK 全チェックOK")

    print("\n=== 当番回数集計（20日まで個人分）===")
    print(f"{'名前':12} {'班':3} {'午前窓':>5} {'午後窓':>5} {'電話':>5} {'窓計':>5}")
    print("-" * 50)
    for name, group in sorted(STAFF, key=lambda x: x[1]):
        print(f"{name:12} {group:3} {am_count[name]:>5} {pm_count[name]:>5} {phone_count[name]:>5} {am_count[name]+pm_count[name]:>5}")

    print("\n=== グループ窓当番（23日以降）===")
    group_idx = 0
    for day in BUSINESS_DAYS:
        if day.day >= GROUP_WINDOW_START_DAY:
            g = GROUP_ORDER[group_idx % len(GROUP_ORDER)]
            wd = WEEKDAY_JP[day.weekday()]
            print(f"  {day.month}/{day.day}({wd}): {g}班")
            group_idx += 1

    output_path = export_to_excel(schedule, am_count, pm_count, phone_count)
    print(f"\n保存完了: {output_path}")
